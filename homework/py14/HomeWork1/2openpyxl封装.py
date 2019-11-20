"""
============================
作者:seak
时间:
邮件:274882401@qq.com
作用：
题目：
============================

为什么要做封装？
为了使用更加方便，提高代码的重用率。

封装的需求是什么？

1、读取数据的方法

2、写入数据的方法
"""
#第一步：
#定义一个读取EXCLE表格的类
#补充解释ReadExcel类继承object对象，拥有了好多可操作对象，这些都是类中的高级特性，py3中不写也会默认拥有
import openpyxl

#定义一个读取EXCLE表格的类
from homework.py14.HomeWork1.readexcel import ReadExcel


class ReadExcel(object):
    #__init__（self）方法中，只有一个self，指的是实例的本身，但是在方法的类部，包含两个属性，name， score下面的这个即是在定义方法时，就直接给定了两个参数，
    #__init__(self,filename,sheet_name)很明显就是直接实例化时，传入相应的参数，而第一种，则需要实例化之后，对属性进行赋值
    def __init__(self,filename,sheet_name):
        '''
        初始化方法
        :param filename: excle文件名
        :param sheet_name: 表单名
        '''
        self.filename = filename
        self.sheet_name = sheet_name

    #第二步，读取工作薄
    #1.读取前需要先打开
    def open(self):
        """打开工作簿，选中表单"""
        self.wb = openpyxl.load_workbook(self.filename)#打开excle名
        self.sh = self.wb[self.sheet_name] #sheet_name表单名

    def save(self):
        """保存工作簿对象的方法"""
        self.wb.save(self.filename)
        self.wb.close()   # 释放内存，这一行加不加关系不大，加了可以释放内存

    def read_data(self):
        "读取数据"
        #1.打开工作薄，选中表单
        self.open()
        #获取最大的行
        max_row = self.sh.max_row

        # 读取所有的数据，放到一个列表中
        list_data = []
        for i in range(1, max_row + 1):
            data1 = self.sh.cell(row=i, column=1).value
            data2 = self.sh.cell(row=i, column=2).value
            data3 = self.sh.cell(row=i, column=3).value
            data4 = self.sh.cell(row=i, column=4).value
            list_data.append([data1, data2, data3, data4])
        print(list_data)

        # 创建一个字典，用来存储所有的用来数据
        cases = []
        # 获取第一行数据中的表头
        title = list_data[0]
        for data in list_data[1:]:
            # 遍历第一行之外的其他行数据，聚合打包成字典
            case = dict(zip(title,data))
            cases.append(case)
        return cases

    def write_data(self):
        pass

    excel = ReadExcel("case.xlsx", "login")

    cases = excel.read_data()
    print(cases)
'''
[
{'case_id': 1, 'title': '正常登录', 'data': '("python24","lemonban")', 'expected': '{"code":0, "msg":"登录成功"}'}, 
{'case_id': 2, 'title': '密码错误', 'data': '("python24", "lemonban1")', 'expected': '{"code": 1, "msg": "账号或密码不正确"}'}, 
{'case_id': 3, 'title': '账户名错误', 'data': '("python", "lemonban")', 'expected': '{"code": 1, "msg": "账号或密码不正确"}'}, 
{'case_id': 4, 'title': '账号为空', 'data': '(None, "lemonban")', 'expected': '{"code": 1, "msg": "所有参数不能为空"}'}, 
{'case_id': 5, 'title': '密码为空', 'data': '("python24", None)', 'expected': '{"code": 1, "msg": "所有参数不能为空"}'}, 
{'case_id': None, 'title': '账号密码都为空', 'data': '("","")', 'expected': None}
]

'''
