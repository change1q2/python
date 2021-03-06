"""
============================
Author:柠檬班-木森
Time:2019/11/18
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""
import openpyxl


class ReadExcel(object):

    def __init__(self, filename, sheet_name):
        """
        初始化方法
        :param filename: excle文件名
        :param sheet_name: 表单名
        """
        self.filename = filename
        self.sheet_name = sheet_name

    def open(self):
        """打开工作簿，选中表单"""
        self.wb = openpyxl.load_workbook(self.filename)
        self.sh = self.wb[self.sheet_name]

    def save(self):
        """保存工作簿对象的方法"""
        self.wb.save(self.filename)
        # self.wb.close()   # 这一行加不加关系不大，加了可以释放内存

    def read_data(self):
        """读取数据"""
        # 1、打开工作薄，选中表单
        self.open()
        # 获取最大的行
        max_row = self.sh.max_row

        # 读取所有的数据，放到一个列表中
        list_data = []
        for i in range(1, max_row + 1):
            data1 = self.sh.cell(row=i, column=1).value
            data2 = self.sh.cell(row=i, column=2).value
            data3 = self.sh.cell(row=i, column=3).value
            data4 = self.sh.cell(row=i, column=4).value
            list_data.append([data1, data2, data3, data4])
        # 创建一个字典，用来存储所有的用来数据
        cases = []
        # 获取第一行数据中的表头
        title = list_data[0]
        for data in list_data[1:]:
            # 遍历第一行之外的其他行数据，聚合打包成字典
            case = dict(zip(title,data))
            cases.append(case)

        return cases



    def write_data(self):
        pass
